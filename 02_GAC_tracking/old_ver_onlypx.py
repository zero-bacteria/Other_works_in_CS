import os
import openpyxl as px
import win32com.client as win32
from datetime import datetime
from openpyxl.styles import Font,Border,Side,Alignment,PatternFill


# 구별을 위한 key를 만드는 함수
# 기존 PR파일을 대상으로 만드는 것 
def make_key(sh):
    # 기존파일이므로 4부터 시작, 제목 적어줌
    sh.cell(3,23).value = 'KEY'
    # 구별할 수 있는 키를 부여
    for i in range(4, sh.max_row +1):
        # SAP PO가 있는것을 대상으로만 진행
        # 근데 무조건 있을수 밖에 없지 않나?
        if sh.cell(i, 21).value == 'O':
            # Order type + PO + Factory + St.code
            sh.cell(i, 23).value = sh.cell(i,2).value + sh.cell(i,6).value +sh.cell(i,7).value + sh.cell(i,9).value
            # print(sh.cell(i,23).value)
# 가능한 리스크 : PCX에 request되지 않은 것들은 키가 달라질 수 있음
# 즉, 이후에 PCX에 리퀘스트되어 KEY가 달라질수 있음을 의미함
# 아마 아닐듯, 확인해보니 Line Plan Season만 달라짐

# 정보를 담기위한 dict 생성, old와 new모두 공통적으로 사용
def make_dict(sh):
    result = dict()
    for i in range(4, sh.max_row + 1):
        temp_key = sh.cell(i,23).value
        if temp_key:
            result[temp_key] = (sh.cell(i,14).value,sh.cell(i,15).value,i) #(GAC,GAC-49,row)
    return result

# 메일 내용을 만들기 위한 함수
# 결과 파일을 바탕으로 해당 row를 직접 추출, html형태를 만들어줌
def make_subject(rows):
    my_table = ''
    for r in rows:
        my_table += '<tr class="mm align-center">'
        # for c in range(1, ws.max_column +1):
        # 지정한 col만 뽑아오게 되어있음, 따라서 my_col이 필요하며 내용 변경시 같이 변경
        for c in my_col:
            temp = ws.cell(r,c).value
            # colorway를 위한 조건문, 간단하게 표현함
            if c == 10 and temp:
              temp = temp[:7]
              temp = temp.replace('-','')
              my_table += f'<td class="mm">{temp}</td>'
            # 만약에 값이 없을경우 빈칸만 추가함
            elif not temp:
              my_table += f'<td class="mm"> </td>'
            else:
            # 일반적인 case를 위한 조건문, 그대로 값 삽입
              my_table += f'<td class="mm">{temp}</td>'
        # 마지막 태그 닫기
        my_table += '</tr>'
    return my_table

# 메일을 보내는 함수 (개인별로 보낼때 사용)
def sendMail(mail_list : list, date): 
    outlook = win32.Dispatch("Outlook.Application")  #Outlook App 객체 생성
    for send_one in mail_list: #mail_list(이름,메일주소)에 대한 리스트를 반복 
        send_mail = outlook.CreateItem(0x0) # 메일 보내기 창 객체 생성 
        send_mail.To = send_one[0] #send_ond[0]은 메일 주소를 불러옴
        send_mail.Subject = f'[NOTICE] GAC DATE CHANGED_{date}' # 메일 제목
        send_mail.HTMLBody = MY_BODY + send_one[1] + MY_TAIL # 메일 내용, 담아서 들어옴
        attachment = os.path.abspath(r) # 해당 파일 경로
        send_mail.Attachments.Add(attachment) # 파일 첨부
        send_mail.Send() # 메일보내기

# 이름을 직접 입력해서 보내는 함수,
# 일단 default 값과 입력해서 보내는 것으로 나누어서 설계
def send_w_name(body,date):
    print('\n\t보내실 분의 이름을 입력해주세요 (고정값은 0을 입력해주세요) : ', end='')
    # 원하는 이름을 불러옴
    name = input()
    # 0을 입력할 경우 default로 들어옴
    if name == '0':
      send_to = '김영균'
    # 아닐 경우 그대로 이름 전달, outlook에서 한글을 인식하여 보내줌.
    elif name == '1':
      send_to = 'DS_T_COSTING@changshininc.com;VJ_Costing@changshininc.com;JJ_Costing@changshininc.com;'
    else:
      send_to = name
    now = date[0:2] + '.' + date[2:4] + '.' +date[4:6]
    outlook = win32.Dispatch("Outlook.Application") 
    send_mail = outlook.CreateItem(0x0) 
    send_mail.To = send_to # 앞서 정리한 주소로 발송,
    send_mail.Subject = f'[NOTICE] GAC DATE CHANGED_{now}' 
    send_mail.HTMLBody = MY_BODY + body + MY_TAIL 
    attachment = os.path.abspath(r) 
    send_mail.Attachments.Add(attachment)
    send_mail.Send()

# 엑셀 뷰를 위한 함수, 문자 길이에 따라 칸을 설정해줌
def AutoFitColumnSize(worksheet, columns=None, margin=2):
    # 각각 col 숫자와, cell을 불러옴
    for i, column_cells in enumerate(worksheet.columns):
        # 검사를 위한 변수
        is_ok = False

        if columns == None:
            is_ok = True
        elif isinstance(columns, list) and i in columns:
            is_ok = True
        
        if is_ok:
            length = max(len(str(cell.value)) for cell in column_cells)
            worksheet.column_dimensions[column_cells[0].column_letter].width = length + margin

    return worksheet

# 중간에 나간 공지문 글귀
# <p>From this week, This mail inform you not only GAC changing but also the new models that have been added suddenly.</p><p>Among the models being added, only models with GAC-49 dates within 2 weeks are included.</p><br>
                        
                        
                        

# HTML을 하나의 파일에 담아내고자 str형태로 설정
MY_BODY = '''
<!doctype html>
<html>
  <head>
    <meta name="viewport" content="width=device-width, initial-scale=1.0"/>
    <meta http-equiv="Content-Type" content="text/html; charset=UTF-8" />
    <title>Simple Transactional Email</title>
    <style>
      /* -------------------------------------
          GLOBAL RESETS
      ------------------------------------- */
      
      /*All the styling goes here*/
      
      img {
        border: none;
        -ms-interpolation-mode: bicubic;
        max-width: 100%; 
      }

      body {
        background-color: #f6f6f6;
        font-family: sans-serif;
        -webkit-font-smoothing: antialiased;
        font-size: 12.5px;
        line-height: 1.4;
        margin: 0;
        padding: 0;
        -ms-text-size-adjust: 100%;
        -webkit-text-size-adjust: 100%; 
      }

      table {
        border-collapse: separate;
        mso-table-lspace: 0pt;
        mso-table-rspace: 0pt;
        width: 100%; }
        table td {
          font-family: sans-serif;
          font-size: 12.5px;
          vertical-align: top; 
      }

      .mine {
        width: 100%;
        border: 1px solid #444444;
        border-collapse: collapse;
      }
      .mm {
        border: 1px solid #444444;
        border-collapse: collapse;
        background-color: #ffffff;
        text-align: center;
        align-items: center;
        
}
.mh {
        border: 1px solid #444444;
        border-collapse: collapse;
        background-color: #FFFDD0";
}

      /* -------------------------------------
          BODY & CONTAINER
      ------------------------------------- */

      .body {
        background-color: #f6f6f6;
        width: 100%; 
      }

      /* Set a max-width, and make it display as block so it will automatically stretch to that width, but will also shrink down on a phone or something */
      .container {
        display: block;
        margin: 0 auto !important;
        /* makes it centered */
        max-width: 1800px;
        padding: 10px;
        width: 1500px; 
      }

      /* This should also be a block element, so that it will fill 100% of the .container */
      .content {
        box-sizing: border-box;
        display: block;
        margin: 0 auto;
        max-width: 1600px;
        padding: 10px; 
      }

      /* -------------------------------------
          HEADER, FOOTER, MAIN
      ------------------------------------- */
      .main {
        background: #ffffff;
        border-radius: 3px;
        width: 100%; 
      }

      .wrapper {
        box-sizing: border-box;
        padding: 20px; 
      }

      .content-block {
        padding-bottom: 10px;
        padding-top: 10px;
      }

      .footer {
        clear: both;
        margin-top: 10px;
        text-align: center;
        width: 100%; 
      }
        .footer td,
        .footer p,
        .footer span,
        .footer a {
          color: #999999;
          font-size: 12.5px;
          text-align: center; 
      }

      /* -------------------------------------
          TYPOGRAPHY
      ------------------------------------- */
      h1,
      h2,
      h3,
      h4 {
        color: #000000;
        font-family: sans-serif;
        font-weight: 400;
        line-height: 1.4;
        margin: 0;
        margin-bottom: 30px; 
      }

      h1 {
        font-size: 15px;
        font-weight: 300;
        text-align: center;
        text-transform: capitalize; 
      }

      p,
      ul,
      ol {
        font-family: sans-serif;
        font-size: 12.5px;
        font-weight: normal;
        margin: 0;
        margin-bottom: 15px; 
      }
        p li,
        ul li,
        ol li {
          list-style-position: inside;
          margin-left: 5px; 
      }

      a {
        color: #3498db;
        text-decoration: underline; 
      }

      /* -------------------------------------
          BUTTONS
      ------------------------------------- */
      .btn {
        box-sizing: border-box;
        width: 100%; }
        .btn > tbody > tr > td {
          padding-bottom: 15px; }
        .btn table {
          width: auto; 
      }
        .btn table td {
          background-color: #ffffff;
          border-radius: 5px;
          text-align: center; 
      }
        .btn a {
          background-color: #ffffff;
          border: solid 1px #3498db;
          border-radius: 5px;
          box-sizing: border-box;
          color: #3498db;
          cursor: pointer;
          display: inline-block;
          font-size: 12.5px;
          font-weight: bold;
          margin: 0;
          padding: 12px 25px;
          text-decoration: none;
          text-transform: capitalize; 
      }

      .btn-primary table td {
        background-color: #3498db; 
      }

      .btn-primary a {
        background-color: #3498db;
        border-color: #3498db;
        color: #ffffff; 
      }

      /* -------------------------------------
          OTHER STYLES THAT MIGHT BE USEFUL
      ------------------------------------- */
      .last {
        margin-bottom: 0; 
      }

      .first {
        margin-top: 0; 
      }

      .align-center {
        text-align: center; 
      }

      .align-right {
        text-align: right; 
      }

      .align-left {
        text-align: left; 
      }

      .clear {
        clear: both; 
      }

      .mt0 {
        margin-top: 0; 
      }

      .mb0 {
        margin-bottom: 0; 
      }

      .preheader {
        color: transparent;
        display: none;
        height: 0;
        max-height: 0;
        max-width: 0;
        opacity: 0;
        overflow: hidden;
        mso-hide: all;
        visibility: hidden;
        width: 0; 
      }

      .powered-by a {
        text-decoration: none; 
      }

      hr {
        border: 0;
        border-bottom: 1px solid #f6f6f6;
        margin: 20px 0; 
      }

      /* -------------------------------------
          RESPONSIVE AND MOBILE FRIENDLY STYLES
      ------------------------------------- */
      @media only screen and (max-width: 620px) {
        table.body h1 {
          font-size: 13px !important;
          margin-bottom: 10px !important; 
        }
        table.body p,
        table.body ul,
        table.body ol,
        table.body td
        table.body span,
        /* tr.mine th {
            border: 1px solid #000 !important;
        } */
        /* tr.mine td {
            border: 1px solid #000 !important;
        } */
        table.body a {
          font-size: 16px !important; 
        }
        table.body .wrapper,
        table.body .article {
          padding: 10px !important; 
        }
        table.body .content {
          padding: 0 !important; 
        }
        table.body .container {
          padding: 0 !important;
          width: 100% !important; 
        }
        table.body .main {
          border-left-width: 0 !important;
          border-radius: 0 !important;
          border-right-width: 0 !important; 
        }
        table.body .btn table {
          width: 100% !important; 
        }
        table.body .btn a {
          width: 100% !important; 
        }
        table.body .img-responsive {
          height: auto !important;
          max-width: 100% !important;
          width: auto !important; 
        }
      }

      /* -------------------------------------
          PRESERVE THESE STYLES IN THE HEAD
      ------------------------------------- */
      @media all {
        .ExternalClass {
          width: 100%; 
        }
        .ExternalClass,
        .ExternalClass p,
        .ExternalClass span,
        .ExternalClass font,
        .ExternalClass td,
        .ExternalClass div {
          line-height: 100%; 
        }
        .apple-link a {
          color: inherit !important;
          font-family: inherit !important;
          font-size: inherit !important;
          font-weight: inherit !important;
          line-height: inherit !important;
          text-decoration: none !important; 
        }
        #MessageViewBody a {
          color: inherit;
          text-decoration: none;
          font-size: inherit;
          font-family: inherit;
          font-weight: inherit;
          line-height: inherit;
        }
        .btn-primary table td:hover {
          background-color: #34495e !important; 
        }
        .btn-primary a:hover {
          background-color: #34495e !important;
          border-color: #34495e !important; 
        } 
      }

    </style>
  </head>
  <body>
    <table role="presentation" border="0" cellpadding="0" cellspacing="0" class="body">
      <tr>
        <td>&nbsp;</td>
        <td class="container">
          <div class="content">

            <!-- START CENTERED WHITE CONTAINER -->
            <table role="presentation" class="main">

              <!-- START MAIN CONTENT AREA -->
              <tr>
                <td class="wrapper">
                  <table role="presentation" border="0" cellpadding="0" cellspacing="0">
                    <tr>
                      <td>
                        <p>NOTICE : GAC DATE CHANGED </p>
                        <hr></hr>
                                
                        <p>Please be informed that GAC date changed and colorways PO are newly added as below.</p>
                        <p>Tracking your models and do not miss your quote DDD!</p>
                        
                      </td>
                    </tr>
                  </table>
                </td>
              </tr>

            <!-- END MAIN CONTENT AREA -->
            
            <tr>
                <td class="wrapper">
  
                  <table role="presentation" class="main, mine text-align" style="border: 1px solid #000; padding:10px;">

                <!-- START MAIN CONTENT AREA -->
                <tr bgcolor="#FFFDD0"class="mh">
                    <th class="mh">PCC Code</th>
                    <th class="mh">Order Type</th>
                    <th class="mh">Costing Season</th>
                    <th class="mh">PO ID</th>
                    <th class="mh">Factory</th>
                    <th class="mh">Status</th>
                    <th class="mh">Prod. Code</th>
                    <th class="mh">Colorway</th>
                    <th class="mh">Dev. Style Name</th>
                    <th class="mh">TD</th>
                    <th class="mh">DPA</th>
                    <th class="mh">Updated GAC</th>
                    <th class="mh">Updated GAC-49</th>
                    <th class="mh">Previous GAC</th>
                    <th class="mh">Previous GAC-49</th>
                    <th class="mh">Actual PIC</th>
                    <th class="mh">PCX Request</th>
                    <th class="mh">SAP PO</th>
                    <th class="mh">GAC Diff.</th>
                </tr>
                
    
            
            
    '''
MY_TAIL = '''
            </td>
            </tr>
            </table>
              <!-- END MAIN CONTENT AREA -->
              </table>

            <!-- START FOOTER -->
            <div class="footer">
              <table role="presentation" border="0" cellpadding="0" cellspacing="0">
                <tr>
                  <td class="content-block">
                    <span class="apple-link">Changshin Inc, 242, Jangpyeong-ro, Saha-gu, Busan, Republic of Korea</span>
                  </td>
                </tr>
                <tr>
                    <td class="content-block">
                      <br> This is an automatically generated message from DS Costing.
                      <br> Please Email me if there is any question or error.
                      <br> noori.lee@changshininc.com / yeonggyun.kim@changshininc.com
                    </td>
                </tr>
                <tr>
                    <td class="content-block powered-by">
                      Powered-by Yeonggyun.Kim(DS).
                    </td>
                </tr>
              </table>
            </div>
            <!-- END FOOTER -->

          </div>
        </td>
        <td>&nbsp;</td>
      </tr>
    </table>
  </body>
</html>
            '''

# 파일리스트를 불러옴
file_list = list()
# pr파일이 들어있는 경로 설정
f_dir = './production_reports'
# 엑셀 파일(매일오는 데이터를 기반으로 제작)
for f in os.listdir(f_dir):
    if 'xlsx' in f:
        file_list.append(f)

# yes 어제 tod 오늘 가장 최신 두개를 비교
# 앞으로 y,t로 표기
yes_f = px.load_workbook(f'{f_dir}/{file_list[-2]}')
tod_f = px.load_workbook(f'{f_dir}/{file_list[-1]}')
ysh = yes_f.active
tsh = tod_f.active

# 각각의 키와 dict를 만들어줌
make_key(ysh)
make_key(tsh)
ydict = make_dict(ysh)
tdict = make_dict(tsh)

# 저장파일을 불러올 경로
my_root = './roots'
wb = px.load_workbook(f'{my_root}/mail_address.xlsx')
ws = wb.active

# 각각의 주소별로 정보를 넣어줄 dict 생성
# PIC의 이름을 기준으로 주소록을과 대조하여 메일 주소 생성,
team_ad = dict()
for i in range(1, ws.max_row+1):
    n = ws.cell(i,1).value
    a = ws.cell(i,2).value
    team_ad[n] = {'mail_ad':a, 'rows':[]}
# print(team_ad)

# 새로 생긴 것들에 대한 row
new_rows = []
# GAC이 다른 전체 row들을 트래킹 하기 위한 변수
total_rows = []
for k in tdict:
    # print(tdict[k][0], ydict[k][0])
    # 만약 GAC날짜가 다르다면
    if k in ydict and tdict[k][0] != ydict[k][0]:
        # 어제오늘의 GAC들을 들고와줌
        tr = tdict[k][2]
        yr = ydict[k][2]
        total_rows.append((tr,yr))
    elif k not in ydict:
        new_rows.append(tdict[k][2])
        # name = tsh.cell(r, 19).value

        # if name in team_ad: 
        #     name = name.lower()
            # team_ad[name]['rows'].append(tdict[k][2])
            # team_ad[name]['gac'] = (ydict[k][0], tdict[k][0])


# 원하는 col만 가져오기 위한 것
my_col = [1,2,5,6,7,8,9,10,11,12,13,14,15,16,17,21,22,23,25]

# hitory를 위한 파일 생성
wb = px.Workbook()
ws = wb.active
ws.title = 'Total Changed GAC List'

# 제목들을 불러와줌
for i in range(1, tsh.max_column+1):
  if i >15:
      ws.cell(1,i+2).value = tsh.cell(1,i).value
  else:
      ws.cell(1,i).value = tsh.cell(1,i).value
    
# 일부 제목을 수정
ws.cell(1,4).value = 'Planning Season'
ws.cell(1,5).value = 'Costing Season'
# ws.insert_cols(16,2)
ws.cell(1,14).value = 'Updated GAC'
ws.cell(1,15).value = 'Updated GAC-49'
ws.cell(1,16).value = 'Previous GAC'
ws.cell(1,17).value = 'Previous GAC-49'
ws.cell(1,25).value = 'GAP'

# 이부분 점검하기

# # 점검을 위한 임시 저장
# tod_f.save('./test.xlsx')

# 날짜 포맷을 저장
datetime_format = "%Y-%m-%d"
# 새로운 파일에 row를 쓰기 위한 변수
n=1
# print(total_rows)


urgent_new = list()
for r in new_rows:
    new_date = datetime.strptime(tsh.cell(r, 15).value, datetime_format)
    tday = datetime.today()
    if (new_date - tday).days < 14:
        print((tday - new_date).days)

        urgent_new.append(r)


for r in urgent_new:
    n+=1
    for c in range(1, tsh.max_column):
        if c > 15:
            ws.cell(n,c+2).value = tsh.cell(r,c).value
        else:
            ws.cell(n,c).value = tsh.cell(r,c).value
    ws.cell(n, ws.max_column-1).value = 'Newly Updated (<14days)'



# 제목을 제외하고 row를 쓰기 시작
# GAC이 다른 정보들만 불러옴
for j in total_rows:
    n += 1
    # 현재 파일에서 정보들을 불러옴
    for i in range(1, tsh.max_column):
      # Pre-GAC을 위해서 16,17열을 비우기 위한 방법
      if i > 15:
        ws.cell(n,i+2).value = tsh.cell(j[0],i).value
      else:
        ws.cell(n,i).value = tsh.cell(j[0],i).value
    # pre-GAC을 적어줌
    ws.cell(n, 16).value = ysh.cell(j[1],14).value
    ws.cell(n, 17).value = ysh.cell(j[1],15).value
    temp1 = datetime.strptime(ws.cell(n, 14).value, datetime_format)
    temp2 = datetime.strptime(ws.cell(n, 16).value, datetime_format)
    
    # print(temp1, temp2)
    # 변경 차를 입력해줌
    ws.cell(n, ws.max_column).value = (temp1-temp2).days






# 디자인 파트

for i in range(1,ws.max_column+1):
    # 제목열의 경우의 설정
    temp = ws.cell(1,i)
    temp.fill = PatternFill(fgColor="FFFDD0", fill_type="solid")
    temp.font = Font(bold=True, name="Calibri")
    temp.border = Border(left=Side(style='thin'), right=Side(style='thin'), top=Side(style='thin'), bottom=Side(style='thin'))
    temp.alignment = Alignment(horizontal="center", vertical="center")
    # 나머지 경우의 설정
    for j in range(2, ws.max_row+1):
        temp = ws.cell(j,i)
        temp.font = Font(name="Calibri")
        temp.alignment = Alignment(horizontal="center", vertical="center")
        if i == 14 or i == 15:
          temp.font = Font(name="calibri", color="D0312D")
        
# 엑셀뷰 정리
AutoFitColumnSize(ws)

# 현재 날짜를 정리
now = str(datetime.now())[2:10].replace('-', '')

# 현재 작업 날짜를 바탕으로 hitory 파일을 저장
wb.save(f'{my_root}/{now}.xlsx')

mail_list = []

# 이름에 따라서 넣기위한 과정, name에다가 넣어줌
for i in range(2, ws.max_row+1):
    name = ws.cell(i, 21).value
    if name and name in team_ad:
        name = name.lower()
        team_ad[name]['rows'].append(i)

# 각각의 모델 담당자별 모델 정보가 담겨있는 row를 저장해줌
for h in team_ad:
    if team_ad[h]['rows']:
        # print(team_ad[h]['mail_ad'], team_ad[h]['rows'])
        mail_list.append((team_ad[h]['mail_ad'], make_subject(team_ad[h]['rows'])))

# 해당 파일 저장
r = f'{my_root}/{now}.xlsx'
# sendMail(mail_list, now)

if total_rows or urgent_new:
  # 담당자 관계없이 모든 row를 불러옴(결과파일)
  all_row = range(2,ws.max_row+1)
  # 모든 row를 바탕으로 내용생성
  all_list = make_subject(all_row)
  # 이름및 default를 바탕으로한 메일 전송
  send_w_name(all_list, now)
else:
  print('\n\tGAC이 변경된 모델이 존재하지 않습니다. 프로그램을 종료합니다.')






        





