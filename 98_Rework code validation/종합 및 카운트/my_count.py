import os
import openpyxl as px

file_list = os.listdir('./')

print(file_list)
target = ''
for f in file_list:
    if 'xlsx' in f:
        target = f
print(target)

target = './최신화.xlsx'

wb = px.load_workbook(f'./{target}')
ws = wb.active

bom_list = ['UPPER Mtl cost', 'BTTM Mtl cost','Packing cost', 'Packing list','UOM', 'FRT TRM', 'Contry ORG ID', 
'Yield', 'DEFECT %', '2P Part name', '2P Part cost', 'BTTM Weight' ]
bol_list = ['PFC Page #','PFC Step #','PFC Process ', 'Count sheet typo', 'Missing vaiants in count sheet', 'Missing OBS', 'OBS typo' ]

tooling_list = ['Tools cost', 'Tools qty', 'Forecast']

document_list = ['Missing', 'File type (PDF/Excel)', 'File name', 'File location']

team_list = ['PFC', 'Costing', 'TD', 'PE', 'TE', 'Yield', 'IE']

my_list = [bom_list, bol_list, tooling_list, document_list, team_list]
print(bom_list)
my_cols = [48,49,50,51,52]
my_dict = dict()

seasons = list()

for i in range(2, ws.max_row + 1):
    if ws.cell(i, 3).value not in my_dict and ws.cell(i,3).value:
        my_dict[ws.cell(i,3).value] = dict()
        seasons.append(ws.cell(i,3).value)

# print(seasons)


# for i in my_cols:
#     print(ws.cell(1,i).value)

# 시즌 돌면서 항목들 초기화 해줌
for s in my_dict:
    for i in my_list:
        for j in i:
            my_dict[s][j] = 0

# print(my_dict)

# 다섯가지 col 중에서 검색 ( 왜 47인지는 나중에)
for i in my_cols:
    # row 끝까지
    for j in range(2, ws.max_row + 1):     
        # 현재 셀의 값
        now = ws.cell(j, i).value
        # 값이 있다면
        if now:
            # 해당 값에서 띄어쓰기 다 없애줌(중간에 섞여 들어오는 것 방지)
            now = now.replace(' ','')
            # 해당 리스트를 만듬
            now_list = now.split(',')
            # 해당리스트를 반복
            for e in now_list:
                # 다섯가지 분류중 조사
                for k in my_list:
                    # 다섯가지 분류에서 종류별 조사
                    for l in k:
                        season = ws.cell(j, 3).value
                        if l.replace(' ','') == e:
                            my_dict[season][l] += 1

# 시즌별 리워크를 살펴보기 위한 딕셔너리
rework_dict = dict()
# 시즌별로 confirm 된 것과 안된 것을 나누어준다.
for s in seasons:
    rework_dict[s] = {'confirmed':{'Yes':0, 'No':0}, 'other':{'Yes':0, 'No':0}}

# 37열 리워크
# 쭉 해당 row들을 돌면서
for j in range(2, ws.max_row+1):
    # costing season 설정
    season = ws.cell(j,3).value
    # 리워크 여부 판정
    rework_cell = ws.cell(j,44).value
    # pcx 상태 살펴보기
    pcx_status = ws.cell(j,37).value
    # 만약 리워크 셀이 있다면 (값이 없는 경우 방지)
    if rework_cell:
        # confirm된 것이면 해당 값(Yes or No)에 해당하는 값을 더해준다.
        # 그외는 그외에 더해준다.
        if pcx_status == 'Confirmed':
            rework_dict[season]['confirmed'][rework_cell] += 1
        else:
            rework_dict[season]['other'][rework_cell] += 1


print(rework_dict)

# 다른 사유들을 정리하기 위한 딕셔너리
other_reasons = dict()
# 시즌별로 초기화를 해준다.
for s in seasons:
    other_reasons[s] = {'total':0}

# 행을 돌면서 검사
for j in range(2, ws.max_row+1):
    # costing season
    season = ws.cell(j,3).value
    # other season cell
    now_cell = ws.cell(j, 53).value
    # 만약 이유가 존재한다면,
    if now_cell:
        # 해당 이유를 딕셔너리에 넣어주고 값도 초기화(0아님)
        if now_cell not in other_reasons[season]:
            other_reasons[season][now_cell] = 1
        # 있는 사유라면 +1
        else:
            other_reasons[season][now_cell] += 1
        # 일단 있다면 전체 숫자에 더해준다.
        other_reasons[season]['total'] += 1


print(other_reasons)
        

# 추가로 col 왜 하나씩 없는지, 리워크 총 개수, other 개수 추가하기




my_b = px.Workbook()
my_s = my_b.active

# 각각 시즌별 rework 요약을 출력하기 위한 과정

# 2행부터 시작
now_r = 2
for season in rework_dict:
    # 열은 항상 고정 밑으로 데이터 쌓아 나감
    now_c = 2
    # 시즌 입력
    my_s.cell(now_r,now_c).value = season
    # 아래로 내려감
    now_r += 1
    # confirm인지 아닌지 
    for co in rework_dict[season]:
        # 제목적고
        my_s.cell(now_r,now_c).value = co
        # 아래로 내려가 y or n 판정
        for yn in rework_dict[season][co]:
            now_r += 1
            # 각각의 y or n과 개수
            my_s.cell(now_r,now_c).value = yn
            my_s.cell(now_r,now_c+1).value = rework_dict[season][co][yn]
        
        # 토탈 입력
        my_s.cell(now_r+1,now_c).value = 'total'
        my_s.cell(now_r+1,now_c+1).value = rework_dict[season][co]['Yes'] + rework_dict[season][co]['No']

        # 다시 돌아가고 옆으로 적기위해 간다.
        now_r -= 2
        now_c += 2
    # 다음 시즌으로 넘어감
    now_r += 5
    
        

    
# 시즌별 사유들을 담기 위한 과정

# 시작 열
now_c = 8
for s in my_dict:
    # 시즌 기입
    my_s.cell(2, now_c).value = s
    now_r = 3
    # 각각의 사유와 값들을 내려가면서 적어줌
    for k,v in my_dict[s].items():
        my_s.cell(now_r, now_c).value = k
        my_s.cell(now_r, now_c+1).value = v
        now_r += 1
    # 그 외 이유들도 적어줌
    my_s.cell(now_r, now_c).value = 'Other Reason'
    my_s.cell(now_r, now_c+1).value = other_reasons[s]['total']
    
    now_c += 4



# 맨 오른쪽에 other reason에 대한 이유들을 통계해줌
now_c = 30
for s in other_reasons:
    my_s.cell(2, now_c).value = s
    now_r = 3
    for k,v in other_reasons[s].items():
        my_s.cell(now_r, now_c).value = k
        my_s.cell(now_r, now_c+1).value = v
        now_r += 1
    
    now_c += 3







my_b.save('./result.xlsx')
    

print(my_dict)
                    


