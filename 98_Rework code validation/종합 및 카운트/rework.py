import win32com.client as win32
import os
import pandas as pd
import openpyxl as px
import seaborn as sns
import matplotlib.pyplot as plt
import datetime
import my_defs as defs
import codecs
from xmlrpc.client import boolean

def make_key(sh, col=0):
    if col == 0:
        col = sh.max_column + 1
    sh.cell(1, col).value = 'MY KEY'
    for i in range(2, sh.max_row + 1):
        if not sh.cell(i,2).value or not sh.cell(i,4).value or not sh.cell(i,7).value:
            continue
        temp_key = sh.cell(i,2).value + sh.cell(i,4).value + str(sh.cell(i,7).value)
        sh.cell(i, col).value = temp_key

def my_join(total_sh, part_sh, diff=0):
    pass

    

error_dict = {'BOM':['UPPER Mtl cost','BTTM Mtl cost','Packing cost','Packing list','UOM',
'FRT TRM','Contry ORG ID','Yield','DEFECT %','2P Part name','2P Part cost','BTTM Weight'],'BOL':['PFC Page #','PFC Step #',
'PFC Process ','Count sheet typo', 'Missing vaiants in count sheet', 'Missing OBS','OBS typo'],
'Tooling Cost':['Tools cost', 'Tools qty', 'Forecast'], 'Document':['Missing', 'File type(PDF/Excel)', 'File name', 'File location'], 'Team':['PFC','Costing', 'TD',' PE', 'TE', 'Yield','IE']}


# 이전까지 키 기준은 Costing 시즌 + 공장 + 생산코드 로 했다.

# 이전 데이터 접근
pre_data = px.load_workbook('./기존/ddd.xlsx')
pre_sh = pre_data.active

make_key(pre_sh)


pre_dict = dict()

for i in range(2, pre_sh.max_row+1):
    now_key = pre_sh.cell(i, pre_sh.max_column).value
    pre_dict[now_key] = dict()
    for j in range(26,32):
        pre_dict[now_key][pre_sh.cell(1,j).value] = pre_sh.cell(i, j).value

print(pre_dict)

# -----------------
# 여기까지는 임시적인 단계 (이후에는 진행하지 않아도 될 단계)


temp = os.listdir('./업뎃')

wb = px.load_workbook(f'./업뎃/{temp[0]}')
ws = wb.active

max_col = 46
ws.cell(1, max_col+1).value = 'MY KEY'
ws.cell(1, max_col+2).value = 'Team'
ws.cell(1, max_col+3).value = 'BOM'
ws.cell(1, max_col+4).value = 'BOL'
ws.cell(1, max_col+5).value = 'Tooling Cost'
ws.cell(1, max_col+6).value = 'Document'
ws.cell(1, max_col+7).value = 'Other'

for i in range(2, ws.max_row+1):
    if not ws.cell(i,3).value or not ws.cell(i,5).value or not ws.cell(i,8).value:
        continue    

    temp = ws.cell(i,3).value + ws.cell(i,5).value + str(ws.cell(i,8).value)
    ws.cell(i, max_col+1).value = temp
    if temp in pre_dict:
        ws.cell(i, max_col+2).value = pre_dict[temp]['Team']
        ws.cell(i, max_col+3).value = pre_dict[temp]['BOM']
        ws.cell(i, max_col+4).value = pre_dict[temp]['BOL']
        ws.cell(i, max_col+5).value = pre_dict[temp]['Tooling cost']
        ws.cell(i, max_col+6).value = pre_dict[temp]['Document']
        ws.cell(i, max_col+7).value = pre_dict[temp]['Other']

wb.save('./raw_data.xlsx')

# result_book = px.Workbook()
# res_sh = result_book.active
# res_sh.title = 'Quote Data'


# for i in range(1, ws.max_column+1):
#     res_sh.cell(1,i).value = ws.cell(1,i).value


# key_set = set()
# n=2
# x = 0
# for i in range(2, ws.max_row+1):
#     if not ws.cell(i,3).value or not ws.cell(i,5).value or not ws.cell(i,8).value:
#         continue    

#     temp = ws.cell(i,3).value + ws.cell(i,5).value + str(ws.cell(i,8).value)
    
#     if temp not in key_set and ws.cell(i,44).value == 'Yes':    
#         key_set.add(temp)
#         ws.cell(i,max_col+1).value = temp

#         for j in range(1, ws.max_column):
#             res_sh.cell(n,j).value = ws.cell(i,j).value
#         n += 1
#     else:
#         x += 1
#         continue

# result_book.save('./test.xlsx')

df = pd.read_excel('./raw_data.xlsx')
df.head(5)

nf = df.drop_duplicates(['MY KEY'])

is_blank = nf['Quote Rework Indicator'] == 'Yes'

blank_df = nf[is_blank & nf['BOM'].isnull() & nf['Team'].isnull() & nf['BOL'].isnull() & nf['Tooling Cost'].isnull() & nf['Document'].isnull() & nf['Other'].isnull()]

blank_df.head()
len(blank_df)

blank_df.to_excel('./blank.xlsx')