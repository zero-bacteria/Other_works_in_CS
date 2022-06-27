import win32com.client as win32
import pandas as pd
import openpyxl as px
import seaborn as sns
import matplotlib.pyplot as plt
import datetime



wb = px.load_workbook('./기존데이터.xlsx')
ws = wb.active

bb = px.load_workbook('./조사분.xlsx')
bs = bb.active


my_dict = dict()
for i in range(2,bs.max_row+1):
    if bs.cell(i,48).value:
        my_dict[bs.cell(i,47).value] = i

for i in range(2, ws.max_row+1):
    temp = ws.cell(i,47).value
    print(temp)
    if temp in my_dict:
        for j in range(48,55):
            ws.cell(i,j).value = bs.cell(my_dict[temp],j).value

wb.save('./최신화.xlsx')


print(my_dict)


        
