import os
from tabnanny import check
import win32com.client as win32
import openpyxl as px
import pandas as pd


wb = px.load_workbook('my_report.xlsx')
ws = wb.active

ws.delete_cols(1)


# 초기에는 dict를 통한 검사 생각
# 하지만 방향 바꿔서 일종의 키 생성

ws.cell(1, 29).value = 'PO_CW_KEY'
for i in range(2, ws.max_row+1):
    now_po = ws.cell(i, 28).value
    now_cw = ws.cell(i, 8).value
    if now_cw:
        ws.cell(i,28).value = now_po + '_' + now_cw

ws.insert_cols(11)
ws.cell(1, 11).value = 'Remain_Type'



# 여기까지 함 중복제거를 한다음에 remain type을 선정하는 것이 유리
wb.save('./before_remain.xlsx')


# # PO별 검사를 위한 set 초기화
# my_check = set()

# for i in range(1, ws.max_row+1):
#     my_check.add(ws.cell(i, 28).value)


# def my_remove(po_season):
#     po_set = set()
#     for i in range(ws.max_row+1, 1, -1):
#         if po_season == ws.cell(i, 27).value:
#             po_set.add(ws.)
        


# 중복제거하는 코드는 시간이 많이 걸려 나중에 하는걸로



# my_check = set()
# n = 0
# for i in range(ws.max_row, 0, -1):
#     now = ws.cell(i,28).value
    
#     n += 1
#     print(n, now)
#     if now in my_check or not now:
#         ws.delete_rows(i)
#     else:
#         my_check.add(now)



    
# print(my_check)

# print (ws.max_row)
# for i in range(ws.max_row+1, 1, -1):
#     now_po = ws.cell(i, 28).value
#     now_cw = ws.cell(i, 9).value
#     if not now_cw:
#         ws.delete_rows(i)
#         continue

#     print(now_cw)
#     if now_cw in my_check[now_po]:
#         ws.delete_rows(i)
#     else:
#         my_check[now_po].append(now_cw)


# print(my_check)
