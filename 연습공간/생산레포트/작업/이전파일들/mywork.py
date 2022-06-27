import openpyxl as px


wb = px.load_workbook('./SUFAHO22 Remain model document management_220406.xlsx')
ws = wb.active

t = ''
for i in range(1, ws.max_column+1):
    t += ws.cell(1, i).value + ','

print(t)