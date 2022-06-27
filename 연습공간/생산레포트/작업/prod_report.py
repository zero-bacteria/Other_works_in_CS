import os
from tabnanny import check
import win32com.client as win32
import openpyxl as px
import pandas as pd
import datetime

# history 관리를 위해 생성일을 변수로 지정해준다. ex)220412
creation_date = datetime.datetime.today()
creation_date = creation_date.strftime('%Y%m%d')[2:]


# 기존의 파일을 불러올 리스트 생성
file_list = list()
# 현재 경로 지정 ( 꼭 절대경로로 할 필요는 없음 나중에 win32에서 사용하려고 사용)
now_dir = os.path.abspath('./01_PCX')
# 해당 경로의 파일들을 불러오는 과정
for f in os.listdir('./01_PCX'):
    # xls 파일만 불러옴 (xls를 포함시키는 과정에서 xlsx가 포함될 수 있어서 다음과 같이 설정)
    if 'xls' in f and 'xlsx' not in f:
        file_list.append(f)

print(f'\t 다음 파일들을 변환하겠습니다 : {file_list}\n')

# 결과 리스트를 작성
result = []
# xls 파일들을 하나하나 불러와 작업을 실행
# excel을 사용하기 위한 application
excel = win32.Dispatch('Excel.Application')
for f in file_list:
    # 현재 경로에 해당 파일을 불러옴
    pb = excel.Workbooks.Open(f'{now_dir}\{f}')
    # xlsx파일로 바꾸어줌
    pb.SaveAs(f'{now_dir}\{f}x', FileFormat = 51) #FileFormat = 51 is for .xlsx extension
    # 닫음
    pb.Close() #FileFormat = 56 is for .xls extension
    # 해당 결과를 파일 리스트에 넣어줌
    result.append(f+'x')
    print(f'\t\t {f} 변환 완료')

# 엑셀 종료(엑셀 열고 닫는것은 반복문에 굳이 필요한가 검토 필요)
excel.Application.Quit()

# _________________

# 내가 새로 저장할 파일 생성
mybook = px.Workbook()
# 시트 접근
my_sh = mybook.active

# 이름을 한번만 넣기위한 변수(스위치)
my_check = 0
# n은 현재 작업중인 row를 표시하기 위함
n = 0
# 해당 경로에 있는 파일들, 즉, 라인시트 파일들을 돈다.
for f in result:
    pb = px.load_workbook(f'./01_PCX/{f}')
    ps = pb.active
    # 시트 이름을 바탕으로 시즌을 추출
    season = pb.sheetnames[0][-4:]
    # 각각 라인플랜시즌 기입해줌
    # 그룹핑이 되어있기 때문에 이를 해제해주기 위해서 거꾸로 탐색 (row가 밀리지 않도록)
    for i in range(ps.max_row, 0, -1):
        # 그룹핑 되어있는것을 만난다면
        if ps.cell(i,1).value:
            ps.delete_rows(i)
        # 아니면 lineplan-season 기입
        else:
            ps.cell(i,1).value = season

    # 첫번째 행에 라인플랜 attribute이름 넣어줌(시트네임 기준임)
    ps.cell(1,1).value = 'lineplan_season'   
    # col 이름들의 소문자 및 공백을 없애줌
    for j in range(2, ps.max_column):
        col = ps.cell(1,j).value
        col = str(col).lower()
        ps.cell(1,j).value = col.replace(' ', '_')

    
    # 해당되는 데이터들을 받아오기위한 작업 시작
    # 이러면 col이 두번 들어가는 현상 발생 (조치해야함)
    for k in range(1, ps.max_row+1):
        # 만약 이미 파일을 한번 넣었고 맨 첫열이라면 넘긴다.
        if k == 1 and my_check == 1:
            continue
        # 들어가면서 1행부터 추가
        n += 1
        for l in range(1, ps.max_column+1):     
            my_sh.cell(n,l).value = ps.cell(k,l).value

    my_check = 1 # 파일을 한번이상 검사했다는 표시

# 중간에 openpyxl과 pandas등을 사용하면서 정리하지 못하여 파일들이 생기게 되는데 이를 위한 폴더생성
# 있으면 생성x 없으면 생성
if creation_date not in os.listdir('./05_Dummies'):
    os.makedirs(f'./05_Dummies/{creation_date}')


# 일단 임시적으로 pcx 데이터 합친것 저장 (날짜에 맞추어서)
mybook.save(f'./05_dummies/{creation_date}/ldf.xlsx')

# pr에 있는 파일 불러옴 (파일하나만 있는것을 가정 맨처음것만 불러온다.)
pr = os.listdir('./02_Production_Report')[0]

print(pr)

rb = px.load_workbook(f'./02_Production_Report/{pr}')
rs = rb.active

# col name을 특수문자를 제거하고 공백을 _로 교체
for i in range(1, rs.max_column+1):
    temp = rs.cell(1,i).value
    temp = temp.replace('/', '')
    temp = temp.replace('.', '')
    rs.cell(1,i).value = temp.replace(' ', '_')

# 임시로 Production Report 파일 저장
rb.save(f'./05_dummies/{creation_date}/rdf.xlsx')


# 이후 Pandas이용해서 데이터 정리
# --------------------

# 이부분은 굳이 필요가없을듯 openpyxl에서 바로 넘길 수 있으면 바로 넘기기
# 각각의 파일들을 불러옴
ldf = pd.read_excel(f'./05_dummies/{creation_date}/ldf.xlsx')
rdf = pd.read_excel(f'./05_dummies/{creation_date}/rdf.xlsx')

# Costing시즌 + 공장 + prodcode를 만들어서 나만의 키를 생성함
# 각각의 파일들을 서식에 맞게 수정하였음

ldf['factory'] = ldf['sourcing_configuration'].str.slice(start=0, stop=2)
ldf['prod_code'] = ldf['style_number'] + ldf['colorways'].str.slice(start=-8, stop=-5)

ldf['my_key'] = ldf['costing_season'] + ldf['factory']+ ldf['prod_code']

rdf['dev_style'] = rdf['Style_Code'].str.slice(start=0, stop=6)
rdf['my_key'] = rdf['Costing_Season'] + rdf['Prod_Fac'] + rdf['Style_Code']

# 이를 left조인을 통해서 PR 즉 SAP PO를 기준으로 넣어줌
# PCX는 prodcode 기준이므로 같은시즌에 주문이 들어오는 경우는 포함 x
# 하지만 PR의 경우 1st, 2nd, 3rd 모두 있기 때문에 이를 확인
# 이 경우 중복제거를 나중에 할텐데 무조건적으로 1st가 먼저 온다고 확신할 수 없다.
# 이를 위한 코드 생성해야함
raw_df = pd.merge(rdf, ldf, on='my_key', how='left')

raw_df.to_excel(f'./05_dummies/{creation_date}/merged_data.xlsx')



# ----------------------

# 필요한 열을 추가하기 위해 colorway 숫자나 문자로만 된것으로 slice 함
raw_df['colorway2'] = raw_df['colorways'].str.slice(start=0, stop=7)
raw_df['colorway2'] = raw_df['colorway2'].str.replace('-', '')

raw_df['colorway2']

# PO를 기준으로 중복제거를 위해 행을 추가해줌
raw_df['my_po'] = raw_df['PO_ID'].str.slice(start=0, stop=4)

my_df = raw_df[['lineplan_season', 'Planning_Season', 'Costing_Season', 'PCC_Code', 'Prod_Fac', 'OBS_Type','MO_ID', 'Style_Code','colorways', 'colorway2', 'Dev_Style','Status','development_team', 'pcc_developer', 'TD', 'GAC', 'GAC-49', 'CBD_ETQ', 'Document_Posting', '5523_in_PCX', 'YIELD','PFC_(Non_trial_cw)', 'PFC_(RFC_trial_cw)', 'CS_BOM_(TP_X)',
       'CS_BOM_(TP_O)', 'pcc_costing', 'quote_state', 'PO_ID', 'my_po','Colorway', 'my_key']]

# 여기까지 raw데이터를 추출
my_df.to_excel(f'./05_dummies/{creation_date}/raw_report.xlsx')


# 만들어진 데이터를 이용하여 조건에 맞게 가공한다.
wb = px.load_workbook(f'./05_dummies/{creation_date}/raw_report.xlsx')
ws = wb.active

# 먼저 pandas로 인해 생성된 col지워줌 (pandas에서도 지울수 있을 것)
ws.delete_cols(1)

# Reamin_type을 생성해줌 
ws.insert_cols(13)
ws.cell(1, 13).value = 'Remain_Type'


# 초기에는 dict를 통한 검사 생각
# 하지만 방향 바꿔서 일종의 키 생성 (많은 시간 소요)

# PO에따른 key를 만들어준다. PO - Style Code
ws.cell(1, 33).value = 'PO_CW_KEY'
for i in range(2, ws.max_row+1):
    now_po = ws.cell(i, 30).value
    now_code = ws.cell(i, 8).value
    if now_code:
        ws.cell(i, 33).value = now_po + '_' + now_code

# APS 오더의 경우에는 Colorway및 다른 Costing Season이 없다.
# Costing 시즌의 경우에는 my_PO? 에서 가져온다
# Colorway의 경우 PR report에서 가져온 CW를 바탕으로 진행한다.
for i in range(1, ws.max_row +1):
    temp_order = ws.cell(i,6).value
    temp_cw = ws.cell(i,31).value
    if temp_order == 'APS':
        ws.cell(i,3).value = ws.cell(i,30).value
        ws.cell(i,10).value = temp_cw[:7].replace('-', '')

# Remain type을 구분하기 위한 함수
# 공장별로 돌리는 것으로 초기에 생성하여 사용중, 수정가능
# 향후 SP23 생산까지 간다면 먹히지 않을 것 이다. why? 단순히 4시즌을 기준으로 했기 때문
def fcty(fcty):
    # 시즌별 집합 생성
    c_sp = set()
    c_su = set()
    c_fa = set()
    c_ho = set()

    # 맨처음 시즌별로 나누어 준다.
    # 이렇게 하는 이유는 돌면서 넣으면 모든 데이터가 반영이 되지 않기 때문
    for i in range(2, ws.max_row+1):
        if ws.cell(i,5).value == fcty:
            temp =  ws.cell(i, 3).value
            s_code = ws.cell(i,8).value
            if temp == 'SP22':
                c_sp.add(s_code)
            elif temp == 'SU22':
                c_su.add(s_code)
            elif temp == 'FA22':
                c_fa.add(s_code)
            elif temp == 'HO22':
                c_ho.add(s_code)
        
    # 조건 별로 나누어준다.
    for i in range(2, ws.max_row+1):
        if ws.cell(i,5).value == fcty:
            # stylecode를 기준으로 진행
            temp = ws.cell(i, 8).value
            # Remain의 경우 
            if ws.cell(i, 12).value == 'Remain':
                # 현재 시즌을 추출
                cseason = ws.cell(i,3).value
                # 현재 시즌이 SP22인 경우 NCF 처음
                if cseason == 'SP22':
                    ws.cell(i, 13).value = 'New'
                # SU22인 경우 SP22에 안했으면 NCF 처음, 했으면 Old
                # 이러한 방식의 반복
                elif cseason == 'SU22':
                    if temp in c_sp:
                        ws.cell(i,13).value = 'Old'
                    else:
                        ws.cell(i,13).value = 'New'
                elif cseason == 'FA22':
                    if temp in c_sp or temp in c_su:
                        ws.cell(i,13).value = 'Old'
                    else:
                        ws.cell(i,13).value = 'New'
                elif cseason == 'HO22':
                    if temp in c_sp or temp in c_su or temp in c_fa:
                        ws.cell(i,13).value = 'Old'
                    else:
                        ws.cell(i,13).value = 'New'
            # 만약 그냥 원래 New라면 New 기입해주기
            elif ws.cell(i, 12).value == 'New':
                ws.cell(i, 13).value = 'New'
# 공장 리스트
fcty_list = ['JJ', 'QD', 'VJ', 'RJ']

for i in fcty_list:
    fcty(i)
                
# Remain 작업 끝


print('이전파일 작업 시작')

# 채워넣기 시작, 이전버전 파일 불러옴
pre_ver = os.listdir('./03_Pre-version')[0]

pre_wb = px.load_workbook(f'./03_Pre-version/{pre_ver}')
pre_ws = pre_wb.active

# 이전버전 TD 담당자와 PCC를 담을 dict
# 공유방에서 수기로 업데이트하는 자료를 불러오기 위함
pre_td = dict()
pre_pcc = dict()

# 이전버전이 없을 경우에만 실행
for i in range(2, pre_ws.max_row +1):
    # 세값이 존재한다면 임시적인 key를 생성 (my_key랑 동일하게 생성하는 것임)
    if pre_ws.cell(i, 3).value and pre_ws.cell(i,5).value and pre_ws.cell(i, 8).value:
        temp_key = pre_ws.cell(i, 3).value + pre_ws.cell(i,5).value + pre_ws.cell(i, 8).value
        # 들어온 적이 없다면 넣어준다.
        if temp_key not in pre_td:
            pre_td[temp_key] = pre_ws.cell(i, 15).value
        if temp_key not in pre_pcc:
            pre_pcc[temp_key] = pre_ws.cell(i, 4).value

print('pre_td', pre_td)
print('pre_pcc', pre_pcc)


# 생산리포트에서도 마찬가지로 불러올 데이터가 있다.
prodf = os.listdir('./04_Prod_File')[0]
prod_wb = px.load_workbook(f'./04_Prod_File/{prodf}')
prod_ws = prod_wb.active

# ETQ 데이터를 불러오는 과정
prod_etq = dict()
for i in range(1, prod_ws.max_row + 1):
    # 마찬가지로 임시적인 키 발급후 값 넣어주기
    if prod_ws.cell(i, 6).value and prod_ws.cell(i, 2).value and prod_ws.cell(i, 3).value:
        temp = prod_ws.cell(i, 6).value
        temp = temp.replace('-', '')
        temp_key2 = prod_ws.cell(i, 2).value + prod_ws.cell(i, 3).value + temp
        if temp_key2 not in prod_etq:
            prod_etq[temp_key2] = prod_ws.cell(i, 19).value

print(prod_etq)

# 만들어진 dict들을 바탕으로 값을 채워넣을 것 이다.
for i in range(2, ws.max_row+1):
    # 이미 만들어진 키를 사용
    ws_key = ws.cell(i, 32).value
    # 각각의 대응되는 값을 넣는다.
    if ws_key in pre_pcc:
        ws.cell(i, 4).value = pre_pcc[ws_key]
    if ws_key in pre_td:
        ws.cell(i, 15).value = pre_td[ws_key]
    # 만약 넣어진 값이 있고 문자형이 아니라면 넣어준다.(문자형으로 넣을경우 반영안될 수도 있음)
    if ws_key in prod_etq and prod_etq[ws_key] and type(prod_etq[ws_key]) is not str:
        # 원하는 형식으로 바꾸어줌
        str_time = prod_etq[ws_key].strftime('%Y-%m-%d')
        print(prod_etq[ws_key])
        print(str_time)
        # ETQ에 해당 값 넣어줌
        ws.cell(i, 19).value = str_time
    else:
        # 별도로 적지 않은 것은 기존의 GAC-49값을 그대로 반영
        ws.cell(i, 19).value = ws.cell(i,18).value

# 작업을 구분하기 위한 Sheet 카피
result_sht = wb.copy_worksheet(ws)


# 작업을 위한 sheet에는 직접 제목을 달아주고 뒤에 수동으로 excel 파일을 가공할 예정 
labels = ['Line plan season', 'PO Season', 'Costing_Season', 'PCC', 'Factory', 'Order Type', 'DPA' ,'Dev.Style' ,'Colorways in PCX', 'Colorway', 'Model Name', 'New/Remain', 'Remain Type', 'Development_Team', 'PCC TD', 'TD Code', 'GAC', 'GAC-49', 'ETQ', 'Document_Posting', '5523_in_PCX', 'YIELD', 'PFC_(Non_trial_cw)', 'PFC_(RFC_trial_cw)', 'CS_BOM_(TP_X)', 'CS_BOM_(TP_O)', 'PCC PIC (Costing)', 'PCX Status', 'PO_ID', 'PO', 'PR_Colorway', 'MY_KEY', 'PO_KEY']

# label들을 적용시켜줌 
for i in range(1, result_sht.max_column +1):
    result_sht.cell(1, i).value = labels[i-1]


# 날짜에따라 결과 파일을 저장
wb.save(f'00_Result_History/{creation_date}_result.xlsx')



# ---------------------------