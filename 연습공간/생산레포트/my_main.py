import win32com.client as win32
import os
import pandas as pd
import openpyxl as px
import seaborn as sns
import matplotlib.pyplot as plt
import datetime
import my_defs as defs
import codecs
from xmlrpc.client import boolean



# 파일 변환하는 과정 -------------------------------------------------

# 개요 : 먼저 xls 파일을 xlsx로 바꾼 다음에 openpyxl로 데이터를 다 합쳐주었다.

# 기존의 파일을 불러올 리스트 생성
file_list = list()
# 현재 경로 지정 ( 꼭 절대경로로 할 필요는 없음 나중에 win32에서 사용하려고 사용)
now_dir = os.path.abspath('./raw_data')
# 해당 경로의 파일들을 불러오는 과정
for f in os.listdir('./raw_data'):
    # xls 파일만 불러옴 (xls를 포함시키는 과정에서 xlsx가 포함될 수 있어서 다음과 같이 설정)
    if 'xls' in f and 'xlsx' not in f:
        file_list.append(f)

print(f'\t 다음 파일들을 변환하겠습니다 : {file_list}\n')

# 결과 리스트를 작성
result = []
# xls 파일들을 하나하나 불러와 작업을 실행
# excel을 사용하기 위한 application
excel = win32.Dispatch('Excel.Application')
for f in file_list:
    # 현재 경로에 해당 파일을 불러옴
    wb = excel.Workbooks.Open(f'{now_dir}\{f}')
    # xlsx파일로 바꾸어줌
    wb.SaveAs(f'{now_dir}\{f}x', FileFormat = 51) #FileFormat = 51 is for .xlsx extension
    # 닫음
    wb.Close() #FileFormat = 56 is for .xls extension
    # 해당 결과를 파일 리스트에 넣어줌
    result.append(f+'x')
    print(f'\t\t {f} 변환 완료')

# 엑셀 종료(엑셀 열고 닫는것은 반복문에 굳이 필요한가 검토 필요)
excel.Application.Quit()

# 워크북을 생성 (new workbook)
nw = px.Workbook()
# 변환된 파일들을 불러옴
for r in result:
    # 파일을 하나하나를 불러옴
    temp = px.load_workbook(f'{now_dir}/{r}')
    # 해당 파일의 시트에 접근
    ts = temp[temp.sheetnames[0]]
    # 결과로 나올 파일에 똑같은 시트 이름을 붙임 (new sheet)
    ns = nw.create_sheet(temp.sheetnames[0])
    # 있는 모든 셀들을 불러옴(동일한 위치에)
    for i in range(1, ts.max_column + 1):
        for j in range(1, ts.max_row + 1):
            ns.cell(j,i).value = ts.cell(j,i).value
# 기존에 생성된 default sheet를 지움
nw.remove(nw['Sheet'])

# 날짜 형식 변환하는 과정
date_col = [22, 23, 24, 25, 30]

for sheet in nw.sheetnames:
    for col in date_col:
        for row in range(4, nw[sheet].max_row + 1):
            now = nw[sheet].cell(row, col).value
            if boolean(now) == True:
                nw[sheet].cell(row, col).value = defs.date_func(now)

# 파일 이름을 짓기위한 과정

# 시즌 리스트 추출
season_list = [s[-4:] for s in nw.sheetnames]
# 순서대로 정렬해줌
season_mm = defs.season_sort(season_list)
# 현재 날짜를 원하는 형식으로 불러옴
now = str(datetime.datetime.now())[2:10].replace('-', '')
# 파일 이름을 지어줌
file_name = f'{now}_{season_mm[0]}-{season_mm[1]}.xlsx'
# 파일이름으로 저장
nw.save(f'./back_data/{file_name}')


# 파일 변환 끝 -------------------------------------------------


# 데이터 정리하는 과정

shts = nw.sheetnames

# 시트들을 각각 조회
for s in shts:
   
    # 시트네임 기준으로 라인플랜시즌을 따옴
    temp = s[-4:]

    # 각각 라인플랜시즌 기입해줌
    for i in range(nw[s].max_row, 0, -1):
        if nw[s].cell(i,1).value:
            nw[s].delete_rows(i)
        else:
            nw[s].cell(i,1).value = temp

    # 첫번째 행에 라인플랜 넣어줌(시트네임 기준임)
    nw[s].cell(1,1).value = 'lineplan_season'   
    # col 이름을 소문자 및 공백을 없애줌
    for j in range(2, nw[s].max_column):
        col = nw[s].cell(1,j).value
        col = str(col).lower()
        nw[s].cell(1,j).value = col.replace(' ', '_')
   
# 파일이름으로 저장
nw.save(f'./data_set/{file_name}')



# pandas 부분----------------------------------------------------------------

# 파일들 불러와서 해당 파일중에 가장 나중에생성된 파일을 지정함
bd_list = os.listdir('./data_set')
xlist = list()

# 파일리스트 중에 해당 엑셀파일을 접수 (숫자로 구분)
for f in bd_list:
    if 'xlsx' in f and f[0] == '2':
        xlist.append(f)

# 가장 최신파일과 이전 파일을 지정
now_file = sorted(xlist, reverse=True)[0]
recent_file = sorted(xlist, reverse=True)[1]

# 시트별로 데이터 프레임 짜서 불러오는 형식으로 일단 후에 걸릴시 csv로 통합하여 로드하자
dfs = pd.read_excel(f'./data_set/{now_file}', sheet_name=None )
odfs = pd.read_excel(f'./data_set/{recent_file}', sheet_name=None )

# 시트별로 나누어진 데이터를 합쳐줌
df = pd.concat(dfs)
odf = pd.concat(odfs)
 
# 둘데이터 프레임중 cost_sheet_id로 데이터를 구분하기 위한 초기화
dfset = set(df['cost_sheet_id'].values)
odfset = set(odf['cost_sheet_id'].values)

# 이전에 없던 모델들을 선별
new_models = dfset - odfset

# 이전에 있던 모델들을 선별
temp = (dfset & odfset)

# 있던모델들의 프레임
tdf = df[df['cost_sheet_id'].isin(temp)]

# milstone이 다른 항목들의 cost_sheet_id를 추출하는 과정 이를위한 리스트 초기화
clist = list()
for c in list(tdf.cost_sheet_id.values):
    # cost_sheet_id가 같은데 milestone이 다르다면 해당 id를 리스트에 추가
    tdf[tdf['cost_sheet_id'] == c].milestone.values != odf[odf['cost_sheet_id'] == c].milestone.values
    clist.append(c)

# milestone이 수정된 데이터 프레임
mmile_df = df[df['cost_sheet_id'].isin(clist)]
# 새로운 모델들의 데이터 프레임
new_df = df[df['cost_sheet_id'].isin(new_models)]

print(mmile_df)
print(new_df)

# 이미지 파일 만들 폴더 생성
img_dir = file_name[:-5]
os.makedirs(f'./pictures/{img_dir}')
img_dir = f'./pictures/{img_dir}'

# bar plot 만들고 저장
bar_plot = defs.my_bar(df)
plt.savefig(f'{img_dir}/bar_plot.png')



# pie plot 시즌별로 만들고 저장
for season in season_list:
    plt.cla()
    defs.my_pie(df, season)
    plt.savefig(f'{img_dir}/{season}_pieplot.png')





# 메일 작성----------------------------------------------------------------

outlook=win32.Dispatch("Outlook.Application")
Txoutlook = outlook.CreateItem(0)

# GetNamespace을 통해서 원하는 네이 공간 형태을 반환합니다.
# MAPI 만 지원합니다. 따라서 위의 영역은 그냥 고정된 영역이라고 보시면 됩니다.

Txoutlook.To = "yeonggyun.kim@changshininc.com"

Txoutlook.CC = ""

Txoutlook.Subject = "This is test mail for automation"

Txoutlook.HTMLBody = defs.my_html()


Txoutlook.Display(True)

# Txoutlook.send()

