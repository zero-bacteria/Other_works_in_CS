import openpyxl as px
import os


sf = os.listdir('./aset')

print(sf)


wb = px.load_workbook(f'./aset/{sf[0]}')
ws = wb.active


print(ws.max_row)

nb = px.Workbook()
ns = nb.active


print(ws.cell(5,5).value[:4])

col_dict = {1:'Line Plan Season', 2:'PO Season', 3:'Costing Season', 4:'PCC', 5: 'Factory', 6:'DPA', 7: 'Dev.Style'}


my = px.load_workbook('./mysample.xlsx')
mys = my.active

print(mys.columns)



# for r in range(5, ws.max_row+1):
#     ns.cell(r-3, 1).value = ws.cell(r,3).value # Lineplan season
#     ns.cell(r-3, 2).value = ws.cell(r,5).value[:4] # PO season
#     ns.cell(r-3, 3).value = ws.cell(r,4).value # Costing season
#     ns.cell(r-3, 4).value = ws.cell(r,1).value # PCC
#     ns.cell(r-3, 5).value = ws.cell(r,6).value # Factory
#     ns.cell(r-3, 6).value = ws.cell(r,12).value # DPA
#     ns.cell(r-3, 7).value = ws.cell(r,8).value # Dev.Style
#     ns.cell(r-3, 8).value = ws.cell(r,9).value # Colorways in PCX
#     # ns.cell(r-3, 9).value = ws.cell(r,6).value # Colorway
#     ns.cell(r-3, 10).value = ws.cell(r,10).value # Model Name
#     ns.cell(r-3, 11).value = ws.cell(r,6).value # New/Reamin
#     # ns.cell(r-3, 12).value = ws.cell(r,6).value # Remain Type
#     # ns.cell(r-3, 13).value = ws.cell(r,6).value # Development Team
#     # ns.cell(r-3, 14).value = ws.cell(r,6).value # PCC TD
#     ns.cell(r-3, 15).value = ws.cell(r,11).value # TD Code
#     ns.cell(r-3, 16).value = ws.cell(r,5).value # PO
#     ns.cell(r-3, 17).value = ws.cell(r,13).value # Current GAC
#     # ns.cell(r-3, 18).value = ws.cell(r,6).value # Previous GAC
#     ns.cell(r-3, 19).value = ws.cell(r,14).value # GAC - 49
#     # ns.cell(r-3, 20).value = ws.cell(r,6).value # ETQ
#     ns.cell(r-3, 21).value = ws.cell(r,16).value # Documnet Posting
#     ns.cell(r-3, 22).value = ws.cell(r,17).value # 5523 in PCX
#     ns.cell(r-3, 23).value = ws.cell(r,18).value # YIELD
#     ns.cell(r-3, 24).value = ws.cell(r,23).value # PFC(Non trial c/v)
#     ns.cell(r-3, 25).value = ws.cell(r,24).value # PFC(RFC trial c/v)
#     ns.cell(r-3, 26).value = ws.cell(r,25).value # CS BOM (T/F)
#     ns.cell(r-3, 27).value = ws.cell(r,26).value # CS BOM (T/P O)
#     ns.cell(r-3, 28).value = ws.cell(r,28).value # PCC PIC (Costing)
#     # ns.cell(r-3, 29).value = ws.cell(r,6).value # PMO Validator
#     ns.cell(r-3, 30).value = ws.cell(r,27).value # Sephiroth Status
#     # ns.cell(r-3, 31).value = ws.cell(r,6).value # PCX Status
#     # ns.cell(r-3, 32).value = ws.cell(r,6).value # Quote Status Date
#     # ns.cell(r-3, 33).value = ws.cell(r,6).value # Quote Missing/Pending  reason
#     # ns.cell(r-3, 34).value = ws.cell(r,6).value # PCX Request
#     # ns.cell(r-3, 35).value = ws.cell(r,6).value # SAP PO
#     # ns.cell(r-3, 36).value = ws.cell(r,6).value # PMO Validator (Linesheet)
#     ns.cell(r-3, 37).value = ws.cell(r,33).value # DPO
#     # ns.cell(r-3, 38).value = ws.cell(r,6).value # APS Order

# nb.save('./test.xlsx')



