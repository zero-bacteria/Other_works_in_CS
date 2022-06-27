import os
import openpyxl as px

# lf = os.listdir('./aset')

# lb = px.load_workbook(f'./aset/{lf[0]}')
# ls = lb.active

sf = px.load_workbook('./mysample.xlsx')
ss = sf.active


rf = os.listdir('./bset')

rb = px.load_workbook(f'./bset/{rf[0]}')
rs = rb.active

rs.insert_cols(1)
rs.insert_cols(8)

rs.cell(1,1).value = 'Line Plan Season'
rs.cell(1,8).value = ' Colorways in PCX'


my_dict = {'col':{}, 'line':{}}
for i in range(3, ss.max_row):
    my_dict['col'][ss.cell(i,9).value] = ss.cell(i,8).value
    temp = ss.cell(i,2).value + '_' + str(ss.cell(i,9).value)
    my_dict['line'][temp] = ss.cell(i,1).value
    
for i in range(2, rs.max_row):
    if rs.cell(i,2).value and rs.cell(i,9).value:
        temp = rs.cell(i,2).value + '_' + str(rs.cell(i,9).value)
        if temp in my_dict['line']:
            rs.cell(i,1).value = my_dict['line'][temp]
    if rs.cell(i,9).value in my_dict['col']:
        rs.cell(i,8).value = my_dict['col'][rs.cell(i,9).value]


rb.save('./hi.xlsx')
