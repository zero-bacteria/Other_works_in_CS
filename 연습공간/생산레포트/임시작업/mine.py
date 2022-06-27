import os
import win32com.client as win32
import openpyxl as px
import pandas as pd


# # 기존의 파일을 불러올 리스트 생성
# file_list = list()
# # 현재 경로 지정 ( 꼭 절대경로로 할 필요는 없음 나중에 win32에서 사용하려고 사용)
# now_dir = os.path.abspath('./pcx')
# # 해당 경로의 파일들을 불러오는 과정
# for f in os.listdir('./pcx'):
#     # xls 파일만 불러옴 (xls를 포함시키는 과정에서 xlsx가 포함될 수 있어서 다음과 같이 설정)
#     if 'xls' in f and 'xlsx' not in f:
#         file_list.append(f)

# print(f'\t 다음 파일들을 변환하겠습니다 : {file_list}\n')

# # 결과 리스트를 작성
# result = []
# # xls 파일들을 하나하나 불러와 작업을 실행
# # excel을 사용하기 위한 application
# excel = win32.Dispatch('Excel.Application')
# for f in file_list:
#     # 현재 경로에 해당 파일을 불러옴
#     wb = excel.Workbooks.Open(f'{now_dir}\{f}')
#     # xlsx파일로 바꾸어줌
#     wb.SaveAs(f'{now_dir}\{f}x', FileFormat = 51) #FileFormat = 51 is for .xlsx extension
#     # 닫음
#     wb.Close() #FileFormat = 56 is for .xls extension
#     # 해당 결과를 파일 리스트에 넣어줌
#     result.append(f+'x')
#     print(f'\t\t {f} 변환 완료')

# # 엑셀 종료(엑셀 열고 닫는것은 반복문에 굳이 필요한가 검토 필요)
# excel.Application.Quit()

# 임시로 생성 나중에는 result로 하면 됨

# 기존의 파일을 불러올 리스트 생성
xlist = list()
# 현재 경로 지정 ( 꼭 절대경로로 할 필요는 없음 나중에 win32에서 사용하려고 사용)
now_dir = os.path.abspath('./pcx')
# 해당 경로의 파일들을 불러오는 과정
for f in os.listdir('./pcx'):
    # xls 파일만 불러옴 (xls를 포함시키는 과정에서 xlsx가 포함될 수 있어서 다음과 같이 설정)
    if 'xlsx' in f:
        xlist.append(f)

# _________________

mybook = px.Workbook()
my_sh = mybook.active

n = 0
for f in xlist:
    wb = px.load_workbook(f'./pcx/{f}')
    ws = wb.active
    season = wb.sheetnames[0][-4:]
    # 각각 라인플랜시즌 기입해줌
    for i in range(ws.max_row, 0, -1):
        if ws.cell(i,1).value:
            ws.delete_rows(i)
        else:
            ws.cell(i,1).value = season

    # 첫번째 행에 라인플랜 넣어줌(시트네임 기준임)
    ws.cell(1,1).value = 'lineplan_season'   
    # col 이름을 소문자 및 공백을 없애줌
    for j in range(2, ws.max_column):
        col = ws.cell(1,j).value
        col = str(col).lower()
        ws.cell(1,j).value = col.replace(' ', '_')


    for k in range(1, ws.max_row+1):
        n += 1
        for l in range(1, ws.max_column+1):     
            my_sh.cell(n,l).value = ws.cell(k,l).value

mybook.save('ldf.csv')

pr = os.listdir('./pr')[0]
rb = px.load_workbook(f'./pr/{pr}')

rs = rb.active
for i in range(1, rs.max_column +1):
    temp = rs.cell(1,i).value
    temp = temp.replace('/', '')
    temp = temp.replace('.', '')
    rs.cell(1,i).value = temp.replace(' ', '_')



rb.save('rdf.csv')

    

    

