import os
import openpyxl as px




targetb = px.load_workbook('./sadf.xlsx')
ts = targetb.active

def fcty(fcty):
    c_sp = set()
    c_su = set()
    c_fa = set()
    c_ho = set()

    for i in range(2, ts.max_row+1):
        if ts.cell(i,5).value == fcty:
            temp =  ts.cell(i, 3).value
            s_code = ts.cell(i,7).value
            if temp == 'SP22':
                c_sp.add(s_code)
            elif temp == 'SU22':
                c_su.add(s_code)
            elif temp == 'FA22':
                c_fa.add(s_code)
            elif temp == 'HO22':
                c_ho.add(s_code)
        

    for i in range(2, ts.max_row+1):
        if ts.cell(i,5).value == fcty:
            temp = ts.cell(i, 7).value
            if ts.cell(i, 11).value == 'Remain':
                cseason = ts.cell(i,3).value
                if cseason == 'SP22':
                    ts.cell(i, 12).value = 'New'
                elif cseason == 'SU22':
                    if temp in c_sp:
                        ts.cell(i,12).value = 'Old'
                    else:
                        ts.cell(i,12).value = 'New'
                elif cseason == 'FA22':
                    if temp in c_sp or temp in c_su:
                        ts.cell(i,12).value = 'Old'
                    else:
                        ts.cell(i,12).value = 'New'
                elif cseason == 'HO22':
                    if temp in c_sp or temp in c_su or temp in c_fa:
                        ts.cell(i,12).value = 'Old'
                    else:
                        ts.cell(i,12).value = 'New'
            elif ts.cell(i, 11).value == 'New':
                ts.cell(i, 12).value = 'New'

fcty_list = ['JJ', 'QD', 'VJ', 'RJ']

for i in fcty_list:
    fcty(i)
                


targetb.save('result3.xlsx')






