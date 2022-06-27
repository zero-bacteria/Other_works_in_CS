import os
import openpyxl as px


bdb = px.load_workbook('./20220406_CBD_LIST.xlsx')
bds = bdb.active

targetb = px.load_workbook('./sample_file.xlsx')
ts = targetb.active

prod_set = set()
for i in range(2,bds.max_row+1):
    if bds.cell(i,6).value:
        temp = bds.cell(i,6).value
        temp = temp.replace('-', '')
        prod_set.add(temp)

print(prod_set)

for i in range(2, ts.max_row+1):
    if ts.cell(i, 11).value == 'Remain':
        
        if ts.cell(i,7).value in prod_set:
            ts.cell(i, 12).value = 'Old'
        else:
            ts.cell(i, 12).value = 'New'


targetb.save('result.xlsx')






