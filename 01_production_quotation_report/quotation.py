import os
import win32com.client as win32
import openpyxl as px
import pandas as pd
import datetime

# season별로 정렬 및 비교 등을 위해서 숫자형태로 표현 해줌
# 각각의 시즌별로 숫자를 넣어줌
def make_float_season(x):
    season = x[0:2]
    year = x[2:]
    if season == 'SP':
        s = 0.1
    elif season == 'SU':
        s = 0.2
    elif season == 'FA':
        s = 0.3
    elif season == 'HO':
        s = 0.4
        
    return float(year) + s

# PO ID역시 숫자로 만듬
# 년도.시즌순서
def make_float_PO_ID(x):
    if x:
        season = x[0:2]
        year = x[2:4]
        season_order = float(x[5]) * 0.01

        if season == 'SP':
            s = 0.1
        elif season == 'SU':
            s = 0.2
        elif season == 'FA':
            s = 0.3
        elif season == 'HO':
            s = 0.4

        return float(year) + s + season_order
    else:
        return 999

#--------------------------------------------------------------------------------------------------------------


# history 관리를 위해 생성일을 변수로 지정해준다. ex)220412
creation_date = datetime.datetime.today()
creation_date = creation_date.strftime('%Y%m%d')[2:]


# 기존의 파일을 불러올 리스트 생성
xls_list = list()
xlsx_list = list()
# 현재 경로 지정 ( 꼭 절대경로로 할 필요는 없음 나중에 win32에서 사용하려고 사용)
now_dir = os.path.abspath('./01_PCX')

print(f'\nxls 파일들을 검사합니다. (PCX file) : {os.listdir(now_dir)}\n')


# 해당 경로의 파일들을 불러오는 과정
for f in os.listdir('./01_PCX'):
    # xls 파일만 불러옴 (xls를 포함시키는 과정에서 xlsx가 포함될 수 있어서 다음과 같이 설정)
    if 'xlsx' in f:
        xlsx_list.append(f)
    elif 'xls' in f:
        xls_list.append(f)
    

# 결과 리스트를 작성
result = []

if not xlsx_list:
    print(f'\t 다음 파일들을 변환하겠습니다 : {xls_list}\n')
    # xls 파일들을 하나하나 불러와 작업을 실행
    # excel을 사용하기 위한 application
    excel = win32.Dispatch('Excel.Application')
    for f in xls_list:
        # 현재 경로에 해당 파일을 불러옴
        pb = excel.Workbooks.Open(f'{now_dir}\{f}')
        # xlsx파일로 바꾸어줌
        pb.SaveAs(f'{now_dir}\{f}x', FileFormat = 51) #FileFormat = 51 is for .xlsx extension
        # 닫음
        pb.Close() #FileFormat = 56 is for .xls extension
        # 해당 결과를 파일 리스트에 넣어줌
        result.append(f+'x')
        print(f'\t\t {f} 변환 완료\n')

    # 엑셀 종료(엑셀 열고 닫는것은 반복문에 굳이 필요한가 검토 필요)
    excel.Application.Quit()
    
else:
    result = xlsx_list

print('xls 검사가 완료되었습니다.\n ')

#--------------------------------------------------------------------------------------------------------------

print('pcx 파일 정형화가 시작됩니다.\n')

pcx_wb = px.Workbook()
pcx_ws = pcx_wb.active

# 이름을 한번만 넣기위한 변수(스위치)
my_check = 0
# n은 현재 작업중인 row를 표시하기 위함
n = 0
# 해당 경로에 있는 파일들, 즉, 라인시트 파일들을 돈다.
for f in result:
    pb = px.load_workbook(f'./01_PCX/{f}')
    ps = pb.active
    # 시트 이름을 바탕으로 시즌을 추출
    season = pb.sheetnames[0][-4:]
    # 각각 라인플랜시즌 기입해줌
    # 그룹핑이 되어있기 때문에 이를 해제해주기 위해서 거꾸로 탐색 (row가 밀리지 않도록)
    for i in range(ps.max_row, 0, -1):
        # 그룹핑 되어있는것을 만난다면
        if ps.cell(i,1).value:
            ps.delete_rows(i)
        # 아니면 lineplan-season 기입
        else:
            ps.cell(i,1).value = season

    # 첫번째 행에 라인플랜 attribute이름 넣어줌(시트네임 기준임)
    ps.cell(1,1).value = 'lineplan_season'   
    # col 이름들의 소문자 및 공백을 없애줌
    for j in range(2, ps.max_column):
        col = ps.cell(1,j).value
        col = str(col).lower()
        ps.cell(1,j).value = col.replace(' ', '_')

    
    # 해당되는 데이터들을 받아오기위한 작업 시작
    # 이러면 col이 두번 들어가는 현상 발생 (조치해야함)
    for k in range(1, ps.max_row+1):
        # 만약 이미 파일을 한번 넣었고 맨 첫열이라면 넘긴다.
        if k == 1 and my_check == 1:
            continue
        # 들어가면서 1행부터 추가
        n += 1
        for l in range(1, ps.max_column+1):     
            pcx_ws.cell(n,l).value = ps.cell(k,l).value

    my_check = 1 # 파일을 한번이상 검사했다는 표시


# pcx 데이터를 불러와서 dataframe으로 바꾸어줌
pcx_df = pd.DataFrame(pcx_ws.values)
# pcx col 및 row를 설정해줌(맨 처음 인덱스가 추가되고 밀려서 나와서 해당 과정 거쳐야 함)
pcx_df.columns = pcx_df.iloc[0,:]
pcx_df = pcx_df.iloc[1:, :]
pcx_df = pcx_df.reset_index(drop=True)
# pcx_df.head()

print('pcx 파일 정형화가 완료되었습니다.\n')

#--------------------------------------------------------------------------------------------------------------

print('Production Report 정리가 시작됩니다.\n')

# Production Report 중에서 열려있거나 임시 파일은 ~220701.xlsx 형태로 나타나는데 이를 배제
pr_list = list() 
for f in os.listdir('./02_Production_Report'):
    if '~' not in f:
        pr_list.append(f)

# pr에 있는 파일 불러옴 (가장 최신것을 불러온다.)
pr = pr_list[-1]
print(f'\n 실행할 가장 최신 파일 : {pr_list[-1]}')

# pr은 col이 3행이나 잡혀있기 때문에 해당을 지워준다.
pr_wb = px.load_workbook(f'./02_Production_Report/{pr}')
pr_ws = pr_wb.active

# 병합되어진 셀을 찾아서 직접 병합을 풀어주는 과정
# 해당 과정을 거쳐야 row가 일괄적으로 지워진다.
merged_list = list()

# 해당 레인지의 좌표를 불러온다.
for m_range in pr_ws.merged_cells.ranges:
    merged_list.append(m_range.coord)
# 해당 좌표를 바탕으로 머지된 셀을 해제해준다.
for m_range in merged_list:
    pr_ws.unmerge_cells(str(m_range))

# 1행에 column 들이 오도록 지워준다.
pr_ws.delete_rows(1, 2)


# pr_wb.save('temp.xlsx')
# print(pr_ws.cell(1,1).value)

# col name을 특수문자를 제거하고 공백을 _로 교체
for i in range(1, pr_ws.max_column+1):
    temp = pr_ws.cell(1, i).value

    temp = temp.replace('/', '')
    temp = temp.replace('.', '')
    pr_ws.cell(1,i).value = temp.replace(' ', '_')

# 데이터 프레임화 시킬때 역시 필요한 과정
# 저장하고 새로 불러와도 되지만 불필요한 저장 방지하기 위해
pr_df = pd.DataFrame(pr_ws.values)
pr_df.columns = pr_df.iloc[0,:]
pr_df = pr_df.iloc[1:, :]
pr_df = pr_df.reset_index(drop=True)
# pr_df.head()



# col_list = pr_df.columns
# ncol_list = list()
# for c in col_list:
#     temp = c.replace(' ', '_')
#     temp = temp.replace('.','_')
#     ncol_list.append(temp)

# pr_df.columns = ncol_list

print('Production Report 데이터화가 완료되었습니다.\n')

#--------------------------------------------------------------------------------------------------------------
print('데이터들의 병합을 시작합니다.\n')

# 각종 셀들을 변형시켜서 직접 key를 만들어 준다. PCX & PR
pcx_df['factory'] = pcx_df['sourcing_configuration'].str.slice(start=0, stop=2)
pcx_df['prod_code'] = pcx_df['style_number'] + pcx_df['colorways'].str.slice(start=-8, stop=-5)
pcx_df['my_key'] = pcx_df['costing_season'] + pcx_df['factory']+ pcx_df['prod_code']

pr_df['Dev_style'] = pr_df['Product_Code'].str.slice(start=0, stop=6)
pr_df['my_key'] = pr_df['Costing'] + pr_df['Prod_Fac'] + pr_df['Product_Code']

# 만든 키를 바탕으로 merge해준다.
raw_df = pd.merge(pr_df, pcx_df, on='my_key', how='left')

print('병합이 완료되었습니다.\n')

#--------------------------------------------------------------------------------------------------------------


print('데이터 불러오기 및 전처리가 시작됩니다.\n')




# APS오더를 분별하기 위한 조건, 코스팅시즌이 존재하지 않고 APS 오더일때
my_condition = raw_df.Costing.isnull() & (raw_df.OBS_Type == 'APS')
# PO를 기준으로 Costing season을 넣어줌
raw_df.loc[my_condition, 'Costing'] = raw_df[my_condition].PO_ID.str.slice(start=0, stop=4)

# 단순히 CW만 있기 위해 만들어줌
raw_df['colorway2'] = raw_df.colorways.str.slice(start=0, stop=7).str.replace('-', '')



# 숫자로 된 costing 시즌 및 PO ID생성
raw_df['float_costing_season'] = raw_df.Costing.apply(make_float_season)
raw_df['float_PO_ID'] = raw_df.PO_ID.apply(make_float_PO_ID)

# NCF를 구별하기 위한 공장가 style code 구분
raw_df['fac_prod_key'] = raw_df.Prod_Fac + raw_df.Product_Code

# 맨처음 오는것을 구분하기 위한 sorting
# NCF를 구분하기 위해 정렬해서 맨첫번째로 나오는 것을 기준으로 해당 시즌은 new 아니면 old로 함
# 이때 PO가 늦으나 GAC이 더 빠른 경우가 있는데 이럴때는 고민을 해봐야한다.
# EX) HO22 1st (6월 1일) FA22 3rd(6월 3일)같은 경우는 HO22는 Remain으로 잡혀버리는데 (PO기준이므로) 이때를 고려해봐야한다.
raw_df = raw_df.sort_values(by=['CGAC', 'float_PO_ID', 'Costing'])

# 다 만들어질 시 아래 키
# raw_df.drop_duplicates(subset='my_key')

# 이전 버전을 불러온다.
pre_ver = os.listdir('./03_Pre-version')[-1]
print(f'지난 리포트 : {pre_ver}')
pre_ver_df = pd.read_excel(f'./03_Pre-version/{pre_ver}')

# raw_df['PFC'] = ''
# 해당 col과 이름을 일치시켜준다(혼동을 피하기 위해, 나중에 origin과 updated로 자동으로 구분되기 위해)
pre_ver_df.rename(columns = {'PCC TD':'pcc_developer'},inplace=True)

# print(pre_ver_df.columns)

# 필요한 col만 merge시킨다 이때 중복인 경우 origin과 updated로 나누어 준다.
raw_df = pd.merge(raw_df, pre_ver_df[['my_key', 'PFC', 'pcc_developer']], how='left', on='my_key', suffixes=('_orgin', '_updated'))

raw_df['pcc_developer'] = raw_df.pcc_developer_updated
# raw_df['PFC'] = raw_df.PFC_updated

# 가장 최신 생산 파일을 불러온다.
prodf = os.listdir('./04_Prod_File')[-1]
prod_file_df = pd.read_excel(f'./04_Prod_File/{prodf}')
print(f'지난 생산 파일 : {prodf}\n')

# 칼럼들을 정리해준다.
col_list = prod_file_df.columns

ncol_list = list()
for c in col_list:
    temp = c.replace(' ', '_')
    temp = temp.replace('.','_')
    ncol_list.append(temp)

prod_file_df.columns = ncol_list

prod_file_df.rename(columns = {'PCC_PIC_(Costing)':'pcc_costing'},inplace=True)
prod_file_df.rename(columns = {'ETQ':'CBD_ETQ'}, inplace=True)
                                        

# 대응 시킬 수 있도록 키를 생성해준다.
prod_file_df['my_key'] = prod_file_df.Costing_Season + prod_file_df.Factory + prod_file_df.Dev_Style.str.replace('-','')

# 해당 키를 바탕으로 ETQ와 담당자를 머지해준다.
raw_df = pd.merge(raw_df, prod_file_df[['my_key','CBD_ETQ','pcc_costing']] , how= 'left', on = 'my_key', suffixes=('_origin', '_updated'))

# 일단 ETQ와 담당자는 별도 비교업이 생산분 tracking 파일을 우선으로 한다.
raw_df['CBD_ETQ'] = raw_df['CBD_ETQ_updated']
raw_df['pcc_costing'] = raw_df['pcc_costing_updated']

# CBD ETQ를 날짜만 나오게 설정 (시간 길게 안나오게)
raw_df['CBD_ETQ'] = raw_df.CBD_ETQ.dt.strftime('%Y-%m-%d')

# 해당 데이터는 무조건 날짜 순으로 먼저 정렬하고, 다음 PO, Costing season 순으로 정렬
raw_df = raw_df.sort_values(by=['CGAC', 'float_PO_ID', 'float_costing_season'])

# 날짜 제일 빠른것 놔두고 costing 시즌별 하나의 컬러웨이만 남기고 나머지는 드랍시켜준다.
raw_df = raw_df.drop_duplicates(subset='my_key')

# Status가 있다면 New로 해주기 위한 조건 isnull로 변경 가능?
init_condition = (raw_df.Status == 'Remain') | (raw_df.Status == 'New')

# remain old 즉 NCF에서 나간적 있는 TCO 모델을 선정하기 위한 조건
# Remain이고 첫번째 오는 것 을 제외하고는 모두 중복 old(fac_prod_key - 공장별 생산 기준)
remain_old_condition = (raw_df.Status == 'Remain') & raw_df.fac_prod_key.duplicated(keep='first')

# 해당 New와 Old 기입
raw_df['remain_type'] = ''
raw_df.loc[init_condition, 'remain_type'] = 'New'
raw_df.loc[remain_old_condition, 'remain_type'] = 'Old'

# PFC 조건을 위해 DPA와 새로운 status 생성
# DPA가 있는 것들을 추출하여 float화 시켜준다.
raw_df['float_dpa'] = raw_df.loc[~raw_df.DPA.isnull(), 'DPA'].str.slice(start=0, stop=4).apply(make_float_season)
raw_df['my_status'] = raw_df.Status

# DPA와 Costing 시즌이 같은 경우는 New로 해준다.
# why? PO가 역전되어 들어오는 경우에는 날짜순으로 정렬하기 때문에 Remain이 나타날 수 있기 때문
dpa_costing_condition = (raw_df.float_dpa == raw_df.float_costing_season)

raw_df.loc[dpa_costing_condition, 'my_status'] = 'New'


# TD 코드가 F3나 F4가 들어가는 경우는 무조건 Remain
pfc_tdcode_condition = raw_df.TD.str.contains('F3') | raw_df.TD.str.contains('F4')
raw_df.loc[pfc_tdcode_condition, 'my_status'] = 'Remain'

# 기존에 진행했던 모델, 현재 기준은 3시즌 이상 차이나면 무조건 Remain
pfc_transfer_condition = (raw_df.float_costing_season - raw_df.float_dpa > 0.2)

raw_df.loc[pfc_transfer_condition, 'my_status'] = 'Remain'





# raw_df.loc[~(raw_df.CBD_ETQ.isnull()), 'CBD_ETQ'] = raw_df.loc[~(raw_df.CBD_ETQ.isnull()), 'CBD_ETQ'].astype(str).str.slice(start = 0, stop = 10)

# raw_df.remain_type.value_counts()
# raw_df = pd.merge(raw_df, my_df[['my_key', 'remain_type']], how='left', on='my_key')

# raw_df.to_excel('result.xlsx')

print('데이터 가공이 완료되었습니다.\n')

#--------------------------------------------------------------------------------------------------------------

print('Report를 위한 엑셀파일 추출을 시작합니다.\n')

# 레포트를 위한 칼럼 선별
report_df = raw_df[['lineplan_season', 'Planning', 'Costing', 'PCC_Code', 'Prod_Fac', 'OBS_Type','DPA', 'Product_Code','colorways', 'colorway2', 'Dev_Style_Name','my_status','remain_type', 'PFC','development_team', 'pcc_developer', 'TD', 'CGAC', 'GAC-49', 'CBD_ETQ', 'Document_Posting', '5523_in_PCX', 'YIELD','PFC_(Non_trial_cw)', 'PFC_(RFC_trial_cw)', 'CS_BOM_(TP_X)',
       'CS_BOM_(TP_O)', 'pcc_costing', 'quote_state', 'PO_ID','Colorway', 'my_key']]
report_df

# 해당 칼럼의 label들을 바꾸어준다.
labels = ['Line plan season', 'PO Season', 'Costing Season', 'PCC', 'Factory', 'Order Type', 'DPA' ,'Dev Style' ,'Colorways in PCX', 'Colorway', 'Model Name', 'New/Remain', 'Remain Type', 'PFC','Development_Team', 'PCC TD', 'TD Code', 'CGAC', 'GAC-49', 'ETQ', 'Document_Posting', '5523_in_PCX', 'YIELD', 'PFC_(Non_trial_cw)', 'PFC_(RFC_trial_cw)', 'CS_BOM_(TP_X)', 'CS_BOM_(TP_O)', 'PCC PIC (Costing)', 'PCX Status', 'PO_ID', 'PR_Colorway', 'my_key']
len(labels)
report_df.columns = labels

report_df

report_df.to_excel(f'./00_Result_History/Production quotation management_{creation_date}.xlsx', index=False, sheet_name='Summary')

print('작업이 완료되었습니다.\n')



# raw_df.to_excel('result.xlsx', index=False, sheet_name='Summary')

