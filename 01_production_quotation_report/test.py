import pandas as pd
import os
import openpyxl as px


pr = os.listdir('./02_Production_Report')[-1]

# pr은 col이 3행이나 잡혀있기 때문에 해당을 지워준다.
pr_wb = px.load_workbook(f'./02_Production_Report/{pr}')
pr_ws = pr_wb.active
merged_list = list()

for m_range in pr_ws.merged_cells.ranges:
    merged_list.append(m_range.coord)

for m_range in merged_list:
    pr_ws.unmerge_cells(str(m_range))

pr_wb.save('temp.xlsx')

# # pr_ws.unmerge_cells('A1:AG3')
# # pr_ws.delete_rows(1,2)
# pr_wb.save('temp.xlsx')
# print(pr_ws.cell(1,1).value)